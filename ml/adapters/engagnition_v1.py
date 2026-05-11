"""Engagnition v1 dataset adapter.

Reads Engagnition CSVs (Kim et al. 2024, DOI:10.1038/s41597-024-03132-3) and outputs
canonical Polars DataFrames with all data quality fixes applied.

Dataset layout:
  Baseline condition/P01–P19   (E4 only, no annotations, no questionnaires)
  LPE condition/P20–P38        (E4 + annotations + questionnaires)
  HPE condition/P39–P57        (E4 + annotations + questionnaires)

Data quality fixes applied here (confirmed via full population inspection):
  1. SGTime rebase        P29 LPE GSR only — span > 10,000s → SGTime = UnixTime - UnixTime.min()
  2. Extra GSR columns    P39/P40/P41 HPE — usecols=["SGTime","UnixTime","GSR"] always
  3. GSR zero-null        P27/P28/P40/P41/P44/P49/P55 — GSR == 0.0 → null
  4. ST sentinel null     All participants — ST < 20.0°C → null (P55 has -273.15°C rows)
  5. Annotation clipping  LPE/HPE — clip Engagement/Gaze/Performance to GSR SGTime span
  6. Low-conductance flag P27/P49 — logged as warning; data processed as-is
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import openpyxl
import polars as pl

from ml.adapters.base_adapter import BaseAdapter
from ml.schema.canonical_v1 import InterventionRecord

logger = logging.getLogger(__name__)

# Confirmed via full population inspection
_GSR_ZERO_NULL = {"P27", "P28", "P40", "P41", "P44", "P49", "P55"}
_LOW_CONDUCTANCE = {"P27", "P49"}

# Demographics from Engagnition paper Table 1 (DOI:10.1038/s41597-024-03132-3).
_DEMOGRAPHICS: dict[str, dict] = {
    # Baseline P01–P19
    "P01": {"age": 11, "diagnosis": "ASD"},
    "P02": {"age":  9, "diagnosis": "ASD,ID"},
    "P03": {"age": 11, "diagnosis": "ASD"},
    "P04": {"age": 10, "diagnosis": "ASD"},
    "P05": {"age": 10, "diagnosis": "ASD,ID"},
    "P06": {"age": 10, "diagnosis": "ASD"},
    "P07": {"age":  4, "diagnosis": "ASD"},
    "P08": {"age": 13, "diagnosis": "ASD"},
    "P09": {"age": 10, "diagnosis": "ASD,ID"},
    "P10": {"age": 10, "diagnosis": "ASD"},
    "P11": {"age":  7, "diagnosis": "ADHD"},
    "P12": {"age": 11, "diagnosis": "ASD"},
    "P13": {"age": 12, "diagnosis": "ASD,ID"},
    "P14": {"age":  8, "diagnosis": "ASD,ID"},
    "P15": {"age": 10, "diagnosis": "ASD"},
    "P16": {"age":  9, "diagnosis": "ASD,ID"},
    "P17": {"age": 11, "diagnosis": "ASD,ID"},
    "P18": {"age": 12, "diagnosis": "ASD"},
    "P19": {"age":  6, "diagnosis": "ASD,ID"},
    # LPE P20–P38
    "P20": {"age": 11, "diagnosis": "ASD"},
    "P21": {"age":  9, "diagnosis": "ASD"},
    "P22": {"age": 10, "diagnosis": "ASD"},
    "P23": {"age": 11, "diagnosis": "ASD"},
    "P24": {"age": 10, "diagnosis": "ASD"},
    "P25": {"age": 10, "diagnosis": "ASD,ID"},
    "P26": {"age": 10, "diagnosis": "ASD"},
    "P27": {"age": 11, "diagnosis": "ASD,ID"},
    "P28": {"age":  6, "diagnosis": "ASD,ID"},
    "P29": {"age": 13, "diagnosis": "ASD"},
    "P30": {"age": 10, "diagnosis": "ASD,ID"},
    "P31": {"age": 10, "diagnosis": "ASD"},
    "P32": {"age":  7, "diagnosis": "ADHD"},
    "P33": {"age": 11, "diagnosis": "ASD"},
    "P34": {"age": 12, "diagnosis": "ASD,ID"},
    "P35": {"age":  8, "diagnosis": "ASD,ID"},
    "P36": {"age":  9, "diagnosis": "ASD,ID"},
    "P37": {"age": 12, "diagnosis": "ASD"},
    "P38": {"age":  4, "diagnosis": "ASD"},
    # HPE P39–P57
    "P39": {"age": 11, "diagnosis": "ASD"},
    "P40": {"age":  9, "diagnosis": "ASD,ID"},
    "P41": {"age": 11, "diagnosis": "ASD"},
    "P42": {"age": 10, "diagnosis": "ASD"},
    "P43": {"age": 10, "diagnosis": "ASD,ID"},
    "P44": {"age": 10, "diagnosis": "ASD"},
    "P45": {"age": 10, "diagnosis": "ASD,ID"},
    "P46": {"age": 10, "diagnosis": "ASD"},
    "P47": {"age":  8, "diagnosis": "ASD,ID"},
    "P48": {"age": 13, "diagnosis": "ASD"},
    "P49": {"age": 16, "diagnosis": "ASD"},
    "P50": {"age": 12, "diagnosis": "ASD"},
    "P51": {"age": 12, "diagnosis": "ASD,ID"},
    "P52": {"age": 10, "diagnosis": "ASD"},
    "P53": {"age": 10, "diagnosis": "ASD"},
    "P54": {"age": 10, "diagnosis": "ASD,ID"},
    "P55": {"age": 11, "diagnosis": "ASD,ID"},
    "P56": {"age":  7, "diagnosis": "ADHD"},
    "P57": {"age":  8, "diagnosis": "ASD,ID"},
}

_CONDITION_DIRS = {
    "baseline": "Baseline condition",
    "LPE": "LPE condition",
    "HPE": "HPE condition",
}


class EngagnitionV1Adapter(BaseAdapter):
    SOURCE = "engagnition_v1"
    SCHEMA_VERSION = "1.0"

    def __init__(self, raw_dir: Path, config: dict) -> None:
        super().__init__(raw_dir, config)
        self._interventions: Optional[dict] = None
        self._session_durations: Optional[dict] = None
        self._questionnaire: Optional[dict] = None

    # ------------------------------------------------------------------
    # Public interface
    # ------------------------------------------------------------------

    def list_participants(self) -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = []
        for i in range(1, 20):
            pairs.append((f"P{i:02d}", "baseline"))
        for i in range(20, 39):
            pairs.append((f"P{i:02d}", "LPE"))
        for i in range(39, 58):
            pairs.append((f"P{i:02d}", "HPE"))
        return pairs

    def load(
        self, participant_id: str, condition: str
    ) -> tuple[pl.DataFrame, InterventionRecord, dict]:
        self._ensure_global_data()

        cond_dir = self.raw_dir / _CONDITION_DIRS[condition] / participant_id

        gsr_df = self._read_gsr(cond_dir, participant_id, condition)
        st_df = self._read_st(cond_dir, participant_id)
        acc_df = self._read_acc(cond_dir)
        ann_dfs = (
            self._read_annotations(cond_dir, gsr_df) if condition != "baseline" else {}
        )

        meta = self._build_meta(participant_id, condition)
        key = f"{condition}/{participant_id}"
        intervention = self._interventions.get(
            key,
            InterventionRecord(
                participant_id=participant_id,
                condition=condition,
                intervention_type="none",
                discrete_timestamps_s=[],
            ),
        )

        df = self._build_canonical_df(
            participant_id, condition, intervention.intervention_type,
            meta, gsr_df, st_df, acc_df, ann_dfs,
        )
        return df, intervention, meta

    # ------------------------------------------------------------------
    # Signal readers
    # ------------------------------------------------------------------

    def _read_gsr(
        self, cond_dir: Path, participant_id: str, condition: str
    ) -> pl.DataFrame:
        # columns= handles P39/P40/P41 extra unnamed columns (fix 2)
        df = pl.read_csv(
            cond_dir / "E4GsrData.csv",
            columns=["SGTime", "UnixTime", "GSR"],
            schema_overrides={"SGTime": pl.Float64, "UnixTime": pl.Float64, "GSR": pl.Float64},
        ).rename({"SGTime": "sg_time_s", "UnixTime": "unix_time", "GSR": "gsr_us"})

        # Fix 1: P29 LPE — SGTime is Unix epoch, rebase to session-relative
        if participant_id == "P29" and condition == "LPE":
            span = df["sg_time_s"].max() - df["sg_time_s"].min()
            if span > self.config["sgtime_corrupt_threshold_s"]:
                unix_min = df["unix_time"].min()
                logger.info(
                    f"P29: GSR SGTime corrupted (span={span:.0f}s), "
                    f"rebasing: SGTime = UnixTime - {unix_min:.0f}"
                )
                df = df.with_columns(
                    (pl.col("unix_time") - unix_min).alias("sg_time_s")
                )

        # Fix 3: null out artifact zeros for known participants
        if participant_id in _GSR_ZERO_NULL:
            n_zero = (df["gsr_us"] == 0.0).sum()
            if n_zero > 0:
                logger.info(f"{participant_id}: nulling {n_zero} zero GSR readings")
                df = df.with_columns(
                    pl.when(pl.col("gsr_us") == 0.0)
                    .then(None)
                    .otherwise(pl.col("gsr_us"))
                    .alias("gsr_us")
                )

        # Fix 6: flag known low-conductance participants
        if participant_id in _LOW_CONDUCTANCE:
            logger.warning(
                f"{participant_id}: low GSR conductance — expect near-flat phasic component"
            )

        return df

    def _read_st(self, cond_dir: Path, participant_id: str) -> pl.DataFrame:
        df = pl.read_csv(
            cond_dir / "E4TmpData.csv",
            schema_overrides={"SGTime": pl.Float64, "UnixTime": pl.Float64, "ST": pl.Float64},
        ).rename({"SGTime": "sg_time_s", "UnixTime": "unix_time", "ST": "skin_temp_c"})

        # Fix 4: null ST below physiological floor (catches P55 -273.15°C sentinel)
        threshold = self.config["st_null_below_c"]
        n_below = (df["skin_temp_c"] < threshold).sum()
        if n_below > 0:
            logger.info(f"{participant_id}: nulling {n_below} ST rows below {threshold}°C")
            df = df.with_columns(
                pl.when(pl.col("skin_temp_c") < threshold)
                .then(None)
                .otherwise(pl.col("skin_temp_c"))
                .alias("skin_temp_c")
            )

        return df

    def _read_acc(self, cond_dir: Path) -> pl.DataFrame:
        return pl.read_csv(
            cond_dir / "E4AccData.csv",
            schema_overrides={
                "SGTime": pl.Float64,
                "UnixTime": pl.Float64,
                "Acc_X": pl.Float64,
                "Acc_Y": pl.Float64,
                "Acc_Z": pl.Float64,
                "Acc_SVM": pl.Float64,
            },
        ).rename({
            "SGTime": "sg_time_s",
            "UnixTime": "unix_time",
            "Acc_X": "acc_x",
            "Acc_Y": "acc_y",
            "Acc_Z": "acc_z",
            "Acc_SVM": "acc_svm",
        })

    def _read_annotations(
        self, cond_dir: Path, gsr_df: pl.DataFrame
    ) -> dict[str, pl.DataFrame]:
        # Fix 5: clip annotations to GSR SGTime span
        gsr_min = gsr_df["sg_time_s"].min()
        gsr_max = gsr_df["sg_time_s"].max()

        result: dict[str, pl.DataFrame] = {}
        for fname, src_col, tgt_col, dtype in [
            ("EngagementData.csv", "Engagement", "engagement", pl.Int32),
            ("GazeData.csv",       "Gaze",       "gaze",       pl.Int32),
            ("PerformanceData.csv","Performance", "performance",pl.Int32),
        ]:
            df = pl.read_csv(
                cond_dir / fname,
                schema_overrides={"SGTime": pl.Float64, src_col: dtype},
            ).rename({"SGTime": "sg_time_s", src_col: tgt_col})

            before = len(df)
            df = df.filter(
                (pl.col("sg_time_s") >= gsr_min) & (pl.col("sg_time_s") <= gsr_max)
            )
            clipped = before - len(df)
            if clipped > 0:
                logger.debug(f"Clipped {clipped} {tgt_col} rows outside GSR span")

            result[tgt_col] = df

        return result

    # ------------------------------------------------------------------
    # DataFrame assembly
    # ------------------------------------------------------------------

    def _build_canonical_df(
        self,
        pid: str,
        condition: str,
        itype: str,
        meta: dict,
        gsr_df: pl.DataFrame,
        st_df: pl.DataFrame,
        acc_df: pl.DataFrame,
        ann_dfs: dict[str, pl.DataFrame],
    ) -> pl.DataFrame:
        common = [
            pl.lit(pid).alias("participant_id"),
            pl.lit(self.SOURCE).alias("source_dataset"),
            pl.lit(self.SCHEMA_VERSION).alias("schema_version"),
            pl.lit(condition).alias("condition"),
            pl.lit(itype if condition != "baseline" else None).alias("intervention_type"),
            pl.lit(meta["age"], dtype=pl.Int32).alias("age"),
            pl.lit(meta["diagnosis"]).alias("diagnosis"),
            pl.lit(meta["nasa_tlx_weighted"], dtype=pl.Float64).alias("nasa_tlx_weighted"),
            pl.lit(meta["sus_score"], dtype=pl.Float64).alias("sus_score"),
        ]

        frames: list[pl.DataFrame] = [
            gsr_df.with_columns(common),
            st_df.with_columns(common),
            acc_df.with_columns(common),
        ]
        for df in ann_dfs.values():
            frames.append(df.with_columns(common))

        # diagonal fills missing columns with null across all frames
        return pl.concat(frames, how="diagonal")

    # ------------------------------------------------------------------
    # Metadata helpers
    # ------------------------------------------------------------------

    def _build_meta(self, participant_id: str, condition: str) -> dict:
        demo = _DEMOGRAPHICS.get(participant_id, {})
        duration = self._session_durations.get(f"{condition}/{participant_id}")
        quest = self._questionnaire.get(f"{condition}/{participant_id}", {})
        is_baseline = condition == "baseline"
        return {
            "participant_id": participant_id,
            "condition": condition,
            "age": demo.get("age"),
            "diagnosis": demo.get("diagnosis"),
            "total_session_duration_s": duration,
            "nasa_tlx_weighted": None if is_baseline else quest.get("nasa_tlx_weighted"),
            "sus_score": None if is_baseline else quest.get("sus_score"),
        }

    # ------------------------------------------------------------------
    # XLSX loaders (lazy, cached after first call)
    # ------------------------------------------------------------------

    def _ensure_global_data(self) -> None:
        if self._interventions is None:
            self._load_interventions()
        if self._session_durations is None:
            self._load_session_durations()
        if self._questionnaire is None:
            self._load_questionnaire()

    def _load_interventions(self) -> None:
        """Parse InterventionData.xlsx → {condition/participant_id: InterventionRecord}"""
        path = self.raw_dir / "InterventionData.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        raw: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            # Columns: A(empty), B=participant, C=condition, D=itype_str, E=sg_time_or_str
            if len(row) < 4:
                continue
            pid = row[1]
            cond = row[2]
            itype_str = row[3]
            sg_time_raw = row[4] if len(row) > 4 else None

            if not (pid and cond and itype_str and str(pid).startswith("P")):
                continue

            pid = str(pid).strip()
            cond = str(cond).strip()
            key = f"{cond}/{pid}"

            if "Discrete" in str(itype_str):
                itype = "discrete"
                sg_time: Optional[float] = (
                    float(sg_time_raw)
                    if isinstance(sg_time_raw, (int, float))
                    else None
                )
            elif "Continuous" in str(itype_str):
                itype = "continuous"
                sg_time = None
            else:
                itype = "none"
                sg_time = None

            if key not in raw:
                raw[key] = {"itype": itype, "timestamps": []}
            if itype == "discrete" and sg_time is not None:
                raw[key]["timestamps"].append(sg_time)

        wb.close()

        self._interventions = {
            key: InterventionRecord(
                participant_id=key.split("/", 1)[1],
                condition=key.split("/", 1)[0],
                intervention_type=data["itype"],
                discrete_timestamps_s=data["timestamps"],
            )
            for key, data in raw.items()
        }

    def _load_session_durations(self) -> None:
        """Parse Session Elapsed Time.xlsx → {condition/participant_id: total_seconds}"""
        path = self.raw_dir / "Session Elapsed Time.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        result: dict[str, float] = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            cond = row[0]
            pid = row[1]
            if not (cond and pid and str(pid).startswith("P")):
                continue
            times = [v for v in row[2:] if isinstance(v, (int, float)) and v is not None]
            result[f"{cond}/{pid}"] = float(sum(times))

        wb.close()
        self._session_durations = result

    def _load_questionnaire(self) -> None:
        """Parse Subjective questionnaire.xlsx → {condition/participant_id: {nasa_tlx, sus}}"""
        path = self.raw_dir / "Subjective questionnaire.xlsx"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        result: dict[str, dict] = {}

        # Sheet 1: NASA-TLX — Weighted score is column J (index 9)
        ws1 = wb.worksheets[0]
        for row in ws1.iter_rows(min_row=3, values_only=True):
            cond, pid = row[0], row[1]
            if not (cond and pid and str(pid).startswith("P")):
                continue
            nasa = row[9] if len(row) > 9 else None
            key = f"{cond}/{pid}"
            result[key] = {
                "nasa_tlx_weighted": float(nasa) if isinstance(nasa, (int, float)) else None,
                "sus_score": None,
            }

        # Sheet 2: SUS — Score is column M (index 12)
        ws2 = wb.worksheets[1]
        for row in ws2.iter_rows(min_row=3, values_only=True):
            cond, pid = row[0], row[1]
            if not (cond and pid and str(pid).startswith("P")):
                continue
            sus = row[12] if len(row) > 12 else None
            key = f"{cond}/{pid}"
            if key in result:
                result[key]["sus_score"] = (
                    float(sus) if isinstance(sus, (int, float)) else None
                )

        wb.close()
        self._questionnaire = result
